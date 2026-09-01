import calendar
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_UP
from io import BytesIO

from django.contrib import messages
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook

from .forms import (
    DeseoForm,
    DeudaForm,
    GastoFijoForm,
    GastoForm,
    IngresoForm,
    NotaForm,
    PagoDeudaForm,
    RecordatorioForm,
)
from .models import (
    Deseo,
    Deuda,
    Gasto,
    GastoFijo,
    Ingreso,
    Nota,
    PagoDeuda,
    Recordatorio,
    HistoryEntry,
)


def _rango_semana(hoy: date) -> tuple[date, date]:
    """Lunes y domingo de la semana que contiene `hoy`."""
    lunes = hoy - timedelta(days=hoy.weekday())
    domingo = lunes + timedelta(days=6)
    return lunes, domingo


def _rango_mes(hoy: date) -> tuple[date, date]:
    inicio = hoy.replace(day=1)
    ultimo = hoy.replace(day=calendar.monthrange(hoy.year, hoy.month)[1])
    return inicio, ultimo


def _sumar(qs, campo: str = "monto") -> Decimal:
    return qs.aggregate(total=Sum(campo))["total"] or Decimal("0")


def _ahorro_acumulado() -> Decimal:
    """Dinero que tienes disponible hoy considerando toda tu historia.

    ingresos totales − gastos totales − pagos realizados hacia deudas.
    Las deudas nuevas no reducen el ahorro hasta que se registren pagos.
    """
    ingresos_tot = _sumar(Ingreso.objects.all())
    gastos_tot = _sumar(Gasto.objects.all())
    pagos_deuda_tot = _sumar(PagoDeuda.objects.all())
    return ingresos_tot - gastos_tot - pagos_deuda_tot


def _generar_excel_respuesta(nombre_archivo: str, workbook: Workbook) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}.xlsx"
    return response


def _normalizar_valor(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, datetime):
        if getattr(valor, "tzinfo", None) is not None:
            valor = valor.astimezone(None).replace(tzinfo=None)
        return valor.date() if valor.time() == datetime.min.time() else valor
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return valor
    if valor is None:
        return ""
    return valor


def _agregar_hoja_con_datos(workbook: Workbook, nombre_hoja: str, columnas: list[tuple[str, str]], registros: list[dict]) -> None:
    sheet = workbook.create_sheet(title=nombre_hoja)
    for col_idx, (campo, titulo) in enumerate(columnas, start=1):
        sheet.cell(row=1, column=col_idx, value=titulo)
    for row_idx, registro in enumerate(registros, start=2):
        for col_idx, (campo, _) in enumerate(columnas, start=1):
            valor = _normalizar_valor(registro.get(campo))
            sheet.cell(row=row_idx, column=col_idx, value=valor)


def exportar_excel(request):
    workbook = Workbook()
    workbook.remove(workbook.active)

    _agregar_hoja_con_datos(
        workbook,
        "Resumen",
        [("concepto", "Concepto"), ("valor", "Valor")],
        [
            {"concepto": "Instrucciones", "valor": "Usa esta plantilla para registrar datos nuevos."},
            {"concepto": "No se exportan registros existentes", "valor": "Solo se deja la estructura para cargar información nueva."},
        ],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Ingresos",
        [("fecha", "Fecha"), ("monto", "Monto"), ("fuente", "Fuente"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Gastos",
        [("fecha", "Fecha"), ("monto", "Monto"), ("categoria", "Categoría"), ("descripcion", "Descripción")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "GastosFijos",
        [("nombre", "Nombre"), ("monto", "Monto"), ("dia_pago", "Día de pago"), ("activo", "Activo"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Deudas",
        [("acreedor", "Acreedor"), ("monto_original", "Monto original"), ("fecha", "Fecha"), ("plazo_meses", "Plazo meses"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "PagosDeuda",
        [("deuda_id", "ID deuda"), ("fecha", "Fecha"), ("monto", "Monto"), ("nota", "Nota")],
        [],
    )

    return _generar_excel_respuesta("plantilla_importacion_finanzas", workbook)


def plantilla_importacion(request):
    workbook = Workbook()
    workbook.remove(workbook.active)

    _agregar_hoja_con_datos(
        workbook,
        "Instrucciones",
        [("campo", "Campo"), ("descripcion", "Descripción")],
        [
            {"campo": "tipo", "descripcion": "Ingrese ingreso, gasto, gasto_fijo, deuda o pago_deuda"},
            {"campo": "fecha", "descripcion": "Fecha en formato YYYY-MM-DD"},
            {"campo": "monto", "descripcion": "Monto numérico con decimales"},
            {"campo": "descripcion", "descripcion": "Texto libre para notas o detalle"},
            {"campo": "referencia", "descripcion": "Nombre del acreedor, fuente o categoría"},
        ],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Ingresos",
        [("fecha", "Fecha"), ("monto", "Monto"), ("fuente", "Fuente"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Gastos",
        [("fecha", "Fecha"), ("monto", "Monto"), ("categoria", "Categoría"), ("descripcion", "Descripción")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "GastosFijos",
        [("nombre", "Nombre"), ("monto", "Monto"), ("dia_pago", "Día de pago"), ("activo", "Activo"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "Deudas",
        [("acreedor", "Acreedor"), ("monto_original", "Monto original"), ("fecha", "Fecha"), ("plazo_meses", "Plazo meses"), ("nota", "Nota")],
        [],
    )
    _agregar_hoja_con_datos(
        workbook,
        "PagosDeuda",
        [("deuda_id", "ID deuda"), ("fecha", "Fecha"), ("monto", "Monto"), ("nota", "Nota")],
        [],
    )

    return _generar_excel_respuesta("plantilla_importacion_finanzas", workbook)


def dashboard(request):
    hoy = timezone.localdate()
    lunes, domingo = _rango_semana(hoy)
    inicio_mes, fin_mes = _rango_mes(hoy)

    # Métricas de la semana
    ingresos_semana = _sumar(
        Ingreso.objects.filter(fecha__range=(lunes, domingo))
    )
    gastos_semana = _sumar(
        Gasto.objects.filter(fecha__range=(lunes, domingo))
    )

    # Métricas del mes
    ingresos_mes = _sumar(
        Ingreso.objects.filter(fecha__range=(inicio_mes, fin_mes))
    )
    gastos_mes = _sumar(
        Gasto.objects.filter(fecha__range=(inicio_mes, fin_mes))
    )

    # Gastos fijos del mes
    fijos_activos = GastoFijo.objects.filter(activo=True)
    fijos_total_mes = _sumar(fijos_activos)
    fijos_pendientes_qs = fijos_activos.filter(dia_pago__gte=hoy.day)
    fijos_pendientes_total = _sumar(fijos_pendientes_qs)

    # Disponible para ahorrar en el mes:
    # lo que entra - lo que ya gasté - lo que debo cubrir de fijos aún
    disponible_ahorro = (
        ingresos_mes - gastos_mes - fijos_pendientes_total
    )

    # Disponible en la semana (sin contar fijos, sólo flujo semanal)
    disponible_semana = ingresos_semana - gastos_semana

    # Deudas: en el dashboard solo mostramos las que siguen con saldo > 0.
    # Las liquidadas permanecen en /deudas/ como historial.
    deudas = [d for d in Deuda.objects.all() if not d.liquidada()]
    total_deuda_original = sum(
        (d.monto_original for d in deudas), Decimal("0")
    )
    total_deuda_saldo = sum((d.saldo() for d in deudas), Decimal("0"))
    total_deuda_pagado = total_deuda_original - total_deuda_saldo

    # Ahorro acumulado (histórico) y lista de deseos
    ahorro_acumulado = _ahorro_acumulado()
    deseos_pendientes = list(Deseo.objects.filter(comprado=False)[:5])
    deseos_resumen = [
        {
            "obj": d,
            "alcanza": d.alcanza(ahorro_acumulado),
            "falta": d.falta(ahorro_acumulado),
            "meses": d.meses_para_ahorrar(ahorro_acumulado, disponible_ahorro),
        }
        for d in deseos_pendientes
    ]

    # Chart: ingresos vs gastos por semana del mes actual
    semanas_labels = []
    semanas_ingresos = []
    semanas_gastos = []
    cursor = inicio_mes
    idx = 1
    while cursor <= fin_mes:
        semana_lunes = cursor - timedelta(days=cursor.weekday())
        semana_domingo = semana_lunes + timedelta(days=6)
        rango_inicio = max(semana_lunes, inicio_mes)
        rango_fin = min(semana_domingo, fin_mes)
        semanas_labels.append(f"Sem {idx}")
        semanas_ingresos.append(
            float(
                _sumar(Ingreso.objects.filter(fecha__range=(rango_inicio, rango_fin)))
            )
        )
        semanas_gastos.append(
            float(
                _sumar(Gasto.objects.filter(fecha__range=(rango_inicio, rango_fin)))
            )
        )
        cursor = semana_domingo + timedelta(days=1)
        idx += 1

    recordatorios_pendientes = Recordatorio.objects.filter(
        completado=False, fecha_recordatorio__lte=hoy
    )
    metas_activas = list(Deseo.objects.filter(comprado=False)[:3])
    metas_deudas = [
        {
            "tipo": "deuda",
            "nombre": d.acreedor,
            "monto": d.saldo(),
            "plazo": d.plazo_meses,
            "estado": "Pendiente" if not d.liquidada() else "Liquidada",
        }
        for d in Deuda.objects.filter(monto_original__gt=0)[:3]
    ]

    contexto = {
        "hoy": hoy,
        "lunes": lunes,
        "domingo": domingo,
        "inicio_mes": inicio_mes,
        "fin_mes": fin_mes,
        "ingresos_semana": ingresos_semana,
        "gastos_semana": gastos_semana,
        "disponible_semana": disponible_semana,
        "ingresos_mes": ingresos_mes,
        "gastos_mes": gastos_mes,
        "fijos_total_mes": fijos_total_mes,
        "fijos_pendientes_total": fijos_pendientes_total,
        "fijos_pendientes": fijos_pendientes_qs,
        "fijos_activos": fijos_activos,
        "disponible_ahorro": disponible_ahorro,
        "deudas": deudas,
        "total_deuda_original": total_deuda_original,
        "total_deuda_saldo": total_deuda_saldo,
        "total_deuda_pagado": total_deuda_pagado,
        "ahorro_acumulado": ahorro_acumulado,
        "deseos_resumen": deseos_resumen,
        "recordatorios_pendientes": recordatorios_pendientes,
        "metas_activas": metas_activas,
        "metas_deudas": metas_deudas,
        "ultimos_ingresos": Ingreso.objects.all()[:5],
        "ultimos_gastos": Gasto.objects.all()[:5],
        "chart_labels": json.dumps(semanas_labels),
        "chart_ingresos": json.dumps(semanas_ingresos),
        "chart_gastos": json.dumps(semanas_gastos),
    }
    return render(request, "ahorros/dashboard.html", contexto)


# ---------- Ingresos ----------
def ingresos_lista(request):
    if request.method == "POST":
        form = IngresoForm(request.POST)
        if form.is_valid():
            ingreso = form.save()
            HistoryEntry.objects.create(
                tipo="ingreso",
                referencia=ingreso.fuente,
                monto=ingreso.monto,
                descripcion=ingreso.nota,
                fecha=ingreso.fecha,
            )
            messages.success(request, "Ingreso registrado.")
            return redirect("ahorros:ingresos")
    else:
        form = IngresoForm()
    return render(
        request,
        "ahorros/ingresos.html",
        {"form": form, "objetos": Ingreso.objects.all()},
    )


@require_POST
def ingreso_borrar(request, pk):
    obj = get_object_or_404(Ingreso, pk=pk)
    obj.delete()
    messages.info(request, "Ingreso eliminado.")
    return redirect("ahorros:ingresos")


# ---------- Gastos ----------
def gastos_lista(request):
    if request.method == "POST":
        form = GastoForm(request.POST, request.FILES)
        if form.is_valid():
            gasto = form.save()
            HistoryEntry.objects.create(
                tipo="gasto",
                referencia=gasto.get_categoria_display(),
                monto=gasto.monto,
                descripcion=gasto.descripcion,
                fecha=gasto.fecha,
            )
            messages.success(request, "Gasto registrado.")
            return redirect("ahorros:gastos")
    else:
        form = GastoForm()
    return render(
        request,
        "ahorros/gastos.html",
        {"form": form, "objetos": Gasto.objects.all()},
    )


@require_POST
def gasto_borrar(request, pk):
    obj = get_object_or_404(Gasto, pk=pk)
    obj.delete()
    messages.info(request, "Gasto eliminado.")
    return redirect("ahorros:gastos")


# ---------- Gastos fijos ----------
def fijos_lista(request):
    if request.method == "POST":
        form = GastoFijoForm(request.POST, request.FILES)
        if form.is_valid():
            gasto = form.save()
            HistoryEntry.objects.create(
                tipo="gasto_fijo",
                referencia=gasto.nombre,
                monto=gasto.monto,
                descripcion=gasto.nota,
                fecha=timezone.localdate(),
            )
            messages.success(request, "Gasto fijo guardado.")
            return redirect("ahorros:fijos")
    else:
        form = GastoFijoForm()
    return render(
        request,
        "ahorros/fijos.html",
        {"form": form, "objetos": GastoFijo.objects.all()},
    )


@require_POST
def fijo_borrar(request, pk):
    obj = get_object_or_404(GastoFijo, pk=pk)
    obj.delete()
    messages.info(request, "Gasto fijo eliminado.")
    return redirect("ahorros:fijos")


# ---------- Deudas ----------
def deudas_lista(request):
    if request.method == "POST":
        form = DeudaForm(request.POST, request.FILES)
        if form.is_valid():
            deuda = form.save()
            HistoryEntry.objects.create(
                tipo="deuda",
                referencia=deuda.acreedor,
                monto=deuda.monto_original,
                descripcion=deuda.nota,
                fecha=deuda.fecha,
            )
            messages.success(request, "Deuda registrada.")
            return redirect("ahorros:deuda_detalle", pk=deuda.pk)
    else:
        form = DeudaForm()
    return render(
        request,
        "ahorros/deudas.html",
        {"form": form, "objetos": Deuda.objects.all()},
    )


def deuda_detalle(request, pk):
    deuda = get_object_or_404(Deuda, pk=pk)
    if request.method == "POST":
        form = PagoDeudaForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.deuda = deuda
            pago.save()
            HistoryEntry.objects.create(
                tipo="pago_deuda",
                referencia=deuda.acreedor,
                monto=pago.monto,
                descripcion=pago.nota,
                fecha=pago.fecha,
            )
            messages.success(
                request,
                f"Pago de ${pago.monto} registrado. Saldo actual: ${deuda.saldo()}.",
            )
            return redirect("ahorros:deuda_detalle", pk=deuda.pk)
    else:
        form = PagoDeudaForm()
    return render(
        request,
        "ahorros/deuda_detalle.html",
        {"deuda": deuda, "pagos": deuda.pagos.all(), "form": form},
    )


@require_POST
def deuda_borrar(request, pk):
    obj = get_object_or_404(Deuda, pk=pk)
    obj.delete()
    messages.info(request, "Deuda eliminada.")
    return redirect("ahorros:deudas")


@require_POST
def pago_borrar(request, pk, pago_pk):
    pago = get_object_or_404(PagoDeuda, pk=pago_pk, deuda_id=pk)
    pago.delete()
    messages.info(request, "Pago eliminado.")
    return HttpResponseRedirect(reverse("ahorros:deuda_detalle", args=[pk]))


# ---------- Deseos (lista de deseos / wishlist) ----------
def deseos_lista(request):
    if request.method == "POST":
        form = DeseoForm(request.POST, request.FILES)
        if form.is_valid():
            deseo = form.save()
            HistoryEntry.objects.create(
                tipo="deseo",
                referencia=deseo.nombre,
                monto=deseo.precio,
                descripcion=deseo.nota,
                fecha=timezone.localdate(),
            )
            messages.success(request, "Deseo agregado a la lista.")
            return redirect("ahorros:deseos")
    else:
        form = DeseoForm()

    hoy = timezone.localdate()
    inicio_mes, fin_mes = _rango_mes(hoy)
    ingresos_mes = _sumar(Ingreso.objects.filter(fecha__range=(inicio_mes, fin_mes)))
    gastos_mes = _sumar(Gasto.objects.filter(fecha__range=(inicio_mes, fin_mes)))
    fijos_pendientes_total = _sumar(
        GastoFijo.objects.filter(activo=True, dia_pago__gte=hoy.day)
    )
    disponible_mensual = ingresos_mes - gastos_mes - fijos_pendientes_total

    ahorro = _ahorro_acumulado()
    prioridad_seleccionada = request.GET.get("prioridad", "todas")
    deseos = Deseo.objects.all()
    if prioridad_seleccionada in {"alta", "media", "baja"}:
        deseos = deseos.filter(prioridad=prioridad_seleccionada)

    objetos = []
    for d in deseos:
        if d.comprado:
            progreso = 100
        elif d.precio > 0:
            progreso = int(
                min(
                    100,
                    (ahorro / d.precio * 100).quantize(Decimal("1."), rounding=ROUND_UP),
                )
            )
        else:
            progreso = 0

        progreso = max(0, progreso)
        objetos.append(
            {
                "obj": d,
                "alcanza": d.alcanza(ahorro),
                "falta": d.falta(ahorro),
                "meses": d.meses_para_ahorrar(ahorro, disponible_mensual),
                "progreso": progreso,
            }
        )
    return render(
        request,
        "ahorros/deseos.html",
        {
            "form": form,
            "objetos": objetos,
            "ahorro_acumulado": ahorro,
            "prioridad_seleccionada": prioridad_seleccionada,
        },
    )


@require_POST
def deseo_comprar(request, pk):
    deseo = get_object_or_404(Deseo, pk=pk)
    if deseo.comprado:
        messages.info(request, "Ese deseo ya estaba marcado como comprado.")
        return redirect("ahorros:deseos")

    hoy = timezone.localdate()
    gasto = Gasto.objects.create(
        fecha=hoy,
        monto=deseo.precio,
        categoria="personal",
        descripcion=f"Deseo comprado: {deseo.nombre}",
    )
    deseo.comprado = True
    deseo.fecha_compra = hoy
    deseo.gasto = gasto
    deseo.save()
    HistoryEntry.objects.create(
        tipo="gasto",
        referencia=deseo.nombre,
        monto=gasto.monto,
        descripcion=gasto.descripcion,
        fecha=gasto.fecha,
    )
    messages.success(
        request,
        f"¡{deseo.nombre} marcado como comprado! Se registró un gasto de ${deseo.precio}.",
    )
    return redirect("ahorros:deseos")


@require_POST
def deseo_borrar(request, pk):
    deseo = get_object_or_404(Deseo, pk=pk)
    deseo.delete()
    messages.info(request, "Deseo eliminado.")
    return redirect("ahorros:deseos")


def notas_lista(request):
    if request.method == "POST":
        form = NotaForm(request.POST)
        if form.is_valid():
            nota = form.save()
            HistoryEntry.objects.create(
                tipo="nota",
                referencia=nota.titulo,
                descripcion=nota.contenido[:255],
                fecha=timezone.localdate(),
            )
            messages.success(request, "Nota guardada.")
            return redirect("ahorros:notas")
    else:
        form = NotaForm()
    return render(
        request,
        "ahorros/notas.html",
        {"form": form, "objetos": Nota.objects.all()},
    )


@require_POST
def nota_borrar(request, pk):
    nota = get_object_or_404(Nota, pk=pk)
    nota.delete()
    messages.info(request, "Nota eliminada.")
    return redirect("ahorros:notas")


def recordatorios_lista(request):
    if request.method == "POST":
        form = RecordatorioForm(request.POST)
        if form.is_valid():
            recordatorio = form.save()
            HistoryEntry.objects.create(
                tipo="recordatorio",
                referencia=recordatorio.titulo,
                descripcion=recordatorio.descripcion[:255],
                fecha=recordatorio.fecha_recordatorio,
            )
            messages.success(request, "Recordatorio guardado.")
            return redirect("ahorros:recordatorios")
    else:
        form = RecordatorioForm()
    recordatorios = Recordatorio.objects.all()
    return render(
        request,
        "ahorros/recordatorios.html",
        {"form": form, "objetos": recordatorios},
    )


def agenda(request):
    hoy = timezone.localdate()
    year = request.GET.get("year")
    month = request.GET.get("month")
    try:
        year = int(year) if year else hoy.year
        month = int(month) if month else hoy.month
        fecha_actual = date(year, month, 1)
    except (TypeError, ValueError):
        fecha_actual = hoy.replace(day=1)
        year = fecha_actual.year
        month = fecha_actual.month

    primer_dia_del_mes = fecha_actual
    ultimo_dia_del_mes = fecha_actual.replace(day=calendar.monthrange(year, month)[1])
    mes_anterior = primer_dia_del_mes - timedelta(days=1)
    mes_siguiente = ultimo_dia_del_mes + timedelta(days=1)

    nota_form = NotaForm()
    recordatorio_form = RecordatorioForm()
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        if form_type == "nota":
            nota_form = NotaForm(request.POST)
            if nota_form.is_valid():
                nota = nota_form.save()
                HistoryEntry.objects.create(
                    tipo="nota",
                    referencia=nota.titulo,
                    descripcion=nota.contenido[:255],
                    fecha=nota.creado_en.date(),
                )
                messages.success(request, "Nota guardada en la agenda.")
                return redirect(
                    reverse("ahorros:agenda") + f"?year={year}&month={month}"
                )
        elif form_type == "recordatorio":
            recordatorio_form = RecordatorioForm(request.POST)
            if recordatorio_form.is_valid():
                recordatorio = recordatorio_form.save()
                HistoryEntry.objects.create(
                    tipo="recordatorio",
                    referencia=recordatorio.titulo,
                    descripcion=recordatorio.descripcion[:255],
                    fecha=recordatorio.fecha_recordatorio,
                )
                messages.success(request, "Recordatorio guardado en la agenda.")
                return redirect(
                    reverse("ahorros:agenda") + f"?year={year}&month={month}"
                )

    calendario = calendar.Calendar(firstweekday=0)
    semanas = calendario.monthdatescalendar(year, month)

    recordatorios = Recordatorio.objects.filter(
        fecha_recordatorio__year=year, fecha_recordatorio__month=month
    )
    notas = Nota.objects.filter(
        creado_en__year=year, creado_en__month=month
    )

    eventos = []
    for rec in recordatorios:
        eventos.append(
            {
                "date": rec.fecha_recordatorio,
                "title": rec.titulo,
                "subtitle": rec.descripcion,
                "type": "recordatorio",
                "completed": rec.completado,
            }
        )
    for nota in notas:
        eventos.append(
            {
                "date": nota.creado_en.date(),
                "title": nota.titulo,
                "subtitle": nota.contenido,
                "type": "nota",
                "completed": None,
            }
        )

    proximos = list(
        Recordatorio.objects.filter(
            completado=False, fecha_recordatorio__gte=hoy
        ).order_by("fecha_recordatorio")[:5]
    )

    contexto = {
        "hoy": hoy,
        "year": year,
        "month": month,
        "fecha_actual": fecha_actual,
        "semanas": semanas,
        "eventos": eventos,
        "nota_form": nota_form,
        "recordatorio_form": recordatorio_form,
        "prev_year": mes_anterior.year,
        "prev_month": mes_anterior.month,
        "next_year": mes_siguiente.year,
        "next_month": mes_siguiente.month,
        "proximos": proximos,
        "titulo_mes": fecha_actual.strftime("%B %Y"),
    }
    return render(request, "ahorros/agenda.html", contexto)


@require_POST
def recordatorio_completar(request, pk):
    recordatorio = get_object_or_404(Recordatorio, pk=pk)
    recordatorio.completado = True
    recordatorio.save()
    messages.success(request, "Recordatorio marcado como completado.")
    return redirect("ahorros:recordatorios")


def historial(request):
    lista = HistoryEntry.objects.all().order_by("-fecha", "-id")
    paginator = Paginator(lista, 50)
    page_number = request.GET.get("page")
    try:
        page_obj = paginator.get_page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.get_page(1)
    return render(
        request,
        "ahorros/historial.html",
        {"objetos": page_obj, "page_obj": page_obj},
    )
